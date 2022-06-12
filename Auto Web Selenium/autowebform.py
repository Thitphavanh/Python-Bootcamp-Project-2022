# autoweb.py
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import random
import wikipedia
from openpyxl import load_workbook
import os

photo_path = os.getcwd()
allfile = os.listdir(photo_path)

# --------------DATA--------------
excelfile = load_workbook('fruit.xlsx')
sheet = excelfile.active
# sheet2 = excelfile['Sheet2']
# sheet3 = excelfile[excelfile.sheetnames[3]]
# print(sheet['B1'].value)

count = len(sheet['B'])
# print(count)

fruit_list = []
for i in range(2, count + 1):
	data = sheet.cell(row = i, column = 2).value
	# print(data)
	split = data.split(',')
	if len(split) >= 2:
		# print(split[0])
		if split[0] not in fruit_list:
			fruit_list.append(split[0].strip())
	else:
		split = data.split('(')
		print(split)
		if split[0] not in fruit_list:
			fruit_list.append(split[0].strip())



path = r'C:\Users\Hery\Desktop\Python Bootcamp 2022\Auto Web Selenium\chromedriver_win32\chromedriver.exe'
driver = webdriver.Chrome(path)

url = 'http://www.uncle-machine.com/login'
driver.get(url)

# search = driver.find_element_by_name('q')
# search.send_keys('Manchester United')
# search.send_keys(Keys.RETURN) # press enter

# --------------LOGIN--------------
username = driver.find_element_by_id('username')
username.send_keys('hery2323@gmail.com')

password = driver.find_element_by_id('password')
password.send_keys('123456')

button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
button.click()

# --------------ADD PRODUCT--------------
time.sleep(2)
url2 = 'http://www.uncle-machine.com/addproduct/'
driver.get(url2) # add product page

for f in fruit_list:
	name = driver.find_element_by_id('name')
	name.send_keys(f)

	pc = random.choice([10000, 15500, 16000, 20000 ])
	price = driver.find_element_by_id('price')
	price.send_keys(pc)

	try:
		dt = wikipedia.summary(f)
	except:
		dt = 'Phenomenal Fruit Shop - Contact us for detail'

	detail = driver.find_element_by_id('detail')
	detail.send_keys(dt)

	photo = driver.find_element_by_id('photo')
	if f + '.jpg' in allfile:
		photo_file = os.path.join(photo_path, f + '.jpg')
		photo.send_keys(photo_path)

	elif f + '.png' in allfile:
		photo_file = os.path.join(photo_path, f + '.png')
		photo.send_keys(photo_path)

	button = driver.find_element_by_xpath('/html/body/div[2]/form/button')
	button.click()

# driver.close()