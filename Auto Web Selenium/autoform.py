from openpyxl import load_workbook
import pyautogui as pg
import pyperclip
import random
import wikipedia
import time
import os
import webbrowser

url = 'http://uncle-machine.com/'
webbrowser.open(url)
time.sleep(2)
pg.hotkey('win','up') # Maximize Screen
time.sleep(2)

path = r'C:\Users\Hery\Desktop\Python Bootcamp 2022\Automate Bot'
bt1 = os.path.join(path, 'product.PNG')
# pg.click(bt1)
pg.click('product.PNG')
time.sleep(2)
# pg.click(100, 200)

excelfile = load_workbook('fruit.xlsx')
sheet = excelfile.active
# sheet2 = excelfile['Sheet2']
# sheet3 = excelfile[excelfile.sheetnames[3]]
# print(sheet['B1'].value)

count = len(sheet['B'])
# print(count)

fruit_list = []
for i in range(2, count+1):
	data = sheet.cell(row = i, column = 2).value
	# print(data)
	split = data.split(',')
	if len(split) >= 2:
		# print(split[0])
		if split[0] not in fruit_list:
			fruit_list.append(split[0])
	else:
		split = data.split('(')
		print(split)
		if split[0] not in fruit_list:
			fruit_list.append(split[0])

# for f in fruit_list:
# 	print(f)

# print(fruit_list)
# Work Flow

# ຄລິກບ່ອນເວ໊ບບຣາວເຊີຣ໌
pg.click(100, 200)

for f in fruit_list:
	# ກົດ 6 ຄັ້ງເພື່ອໃຫ້ເຄີເຊີຣຢູ່ໃນບ່ອນໃສ່ຂໍ້ມູນ
	pg.press('tab', presses = 6)

	# ໃສ່ຂໍ້ມູນຊື່ + tab
	product = f
	pyperclip.copy(product)
	pg.hotkey('ctrl','v')
	pg.press('tab')

	# ສຸ່ມລາຄາ
	rand = random.choice(range(100, 1001, 100))

	# ໃສ່ລາຄາ (ໂຕເລກ) + tab
	pyperclip.copy(rand)
	pg.hotkey('ctrl','v')
	pg.press('tab')

	# ຂໍ້ມູນສິນຄ້າ ໃສ່ຊື່ຮ້ານໄປກ່ອນ +  tab
	try: 
		info = wikipedia.summary(f)
	except:
		info = 'No detail - Phenomenal Fruit Shop'
	pyperclip.copy(info)
	pg.hotkey('ctrl','v')
	pg.press('tab')		
							
	# tab ເພື່ອໄປຫາປຸ່ມ submit
	pg.press('tab')
	# enterເພື່ອໄປ summit	100Phenomenal Fruit Shop		
							
	pg.press('enter')
	time.sleep(1)

